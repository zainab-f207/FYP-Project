from typing import Optional, Dict, Any

"""
Report generation utilities for CrimeVision backend.
"""
import os
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, cast
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .core.database import get_db_connection
from .core.config import get_logger

logger = get_logger(__name__)

# Create reports directory if it doesn't exist
REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


def save_report_to_db(title: str, report_type: str, format_type: str, filepath: str, created_by: str, parameters: Optional[Dict[str, Any]] = None) -> int:
    """Save report metadata to the database and return the report ID."""
    conn = get_db_connection()
    cursor = conn.cursor()
    report_id: int = -1
    try:
        if not filepath or not os.path.exists(filepath):
            logger.error(f"save_report_to_db: Invalid or missing file path: {filepath}")
            return -1
        file_size = os.path.getsize(filepath)
        cursor.execute(
            """
            INSERT INTO reports (title, type, format, file_path, file_size, status, created_by, parameters, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (title, report_type, format_type, filepath, file_size, 'completed', created_by, json.dumps(parameters or {}), datetime.now())
        )
        report_id = cursor.lastrowid if cursor.lastrowid is not None else -1
        conn.commit()
    except Exception as e:
        logger.error(f"Error saving report to DB: {e}")
        report_id = -1
    finally:
        cursor.close()
        conn.close()
    return int(report_id)


def generate_crime_summary_pdf(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate PDF report for crime summary with charts, statistics, or table style."""
    import matplotlib.pyplot as plt
    import io
    from reportlab.platypus import Image

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crime_summary_{timestamp}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
    )
    story.append(Paragraph("Crime Summary Report", title_style))
    story.append(Spacer(1, 12))

    # Report info
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 12))

    # Filters applied
    if filters:
        story.append(Paragraph("Filters Applied:", styles['Heading2']))
        for key, value in filters.items():
            if value:
                story.append(Paragraph(f"{key}: {value}", styles['Normal']))
        story.append(Spacer(1, 12))


    # Check if there is any data for the selected date range
    total_crimes = data.get('summary', {}).get('total_crimes', 0)
    if total_crimes == 0:
        story.append(Paragraph("No crime data available for the selected date range.", styles['Heading2']))
        doc.build(story)
        return filepath

    if report_style == "charts":
        # Area distribution chart
        area_distribution = data.get('area_distribution', [])
        if area_distribution:
            areas = [item.get('area', '') for item in area_distribution[:10]]
            counts = [item.get('count', 0) for item in area_distribution[:10]]
            plt.figure(figsize=(6, 4))
            plt.bar(areas, counts)
            plt.title("Crimes by Area (Top 10)")
            plt.xlabel("Area")
            plt.ylabel("Count")
            plt.xticks(rotation=45, ha='right')
            buf = io.BytesIO()
            plt.tight_layout()
            plt.savefig(buf, format='png')
            buf.seek(0)
            story.append(Image(buf))
            plt.close()
            story.append(Spacer(1, 12))

        # Crime type distribution chart
        crime_type_distribution = data.get('crime_type_distribution', [])
        if crime_type_distribution:
            types = [item.get('crime_type', '') for item in crime_type_distribution[:10]]
            counts = [item.get('count', 0) for item in crime_type_distribution[:10]]
            plt.figure(figsize=(6, 4))
            plt.bar(types, counts, color='orange')
            plt.title("Crimes by Type (Top 10)")
            plt.xlabel("Crime Type")
            plt.ylabel("Count")
            plt.xticks(rotation=45, ha='right')
            buf = io.BytesIO()
            plt.tight_layout()
            plt.savefig(buf, format='png')
            buf.seek(0)
            story.append(Image(buf))
            plt.close()
            story.append(Spacer(1, 12))

        # Monthly trend chart
        monthly_trend = data.get('monthly_trend', [])
        if monthly_trend:
            months = [item.get('month', '') for item in monthly_trend]
            counts = [item.get('count', 0) for item in monthly_trend]
            plt.figure(figsize=(7, 4))
            plt.plot(months, counts, marker='o')
            plt.title("Monthly Crime Trend (Last 12 Months)")
            plt.xlabel("Month")
            plt.ylabel("Count")
            plt.xticks(rotation=45, ha='right')
            buf = io.BytesIO()
            plt.tight_layout()
            plt.savefig(buf, format='png')
            buf.seek(0)
            story.append(Image(buf))
            plt.close()
            story.append(Spacer(1, 12))

    elif report_style == "statistics":
        # Summary statistics only
        story.append(Paragraph("Summary Statistics:", styles['Heading2']))
        summary = data.get('summary', {})
        summary_data = [
            ["Metric", "Value"],
            ["Total Crimes", str(summary.get('total_crimes', 0))],
            ["Unique Areas", str(summary.get('unique_areas', 0))],
            ["Unique Crime Types", str(summary.get('unique_crime_types', 0))],
            ["High Risk %", f"{summary.get('high_risk_percentage', 0.0):.1f}%"],
            ["Medium Risk %", f"{summary.get('medium_risk_percentage', 0.0):.1f}%"],
            ["Low Risk %", f"{summary.get('low_risk_percentage', 0.0):.1f}%"],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

    else:  # table (raw data)
        # Summary statistics
        story.append(Paragraph("Summary Statistics:", styles['Heading2']))
        summary = data.get('summary', {})
        summary_data = [
            ["Metric", "Value"],
            ["Total Crimes", str(summary.get('total_crimes', 0))],
            ["Unique Areas", str(summary.get('unique_areas', 0))],
            ["Unique Crime Types", str(summary.get('unique_crime_types', 0))],
            ["High Risk %", f"{summary.get('high_risk_percentage', 0.0):.1f}%"],
            ["Medium Risk %", f"{summary.get('medium_risk_percentage', 0.0):.1f}%"],
            ["Low Risk %", f"{summary.get('low_risk_percentage', 0.0):.1f}%"],
        ]
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Area distribution table
        area_distribution = data.get('area_distribution', [])
        if area_distribution:
            story.append(Paragraph("Crimes by Area (Top 10):", styles['Heading2']))
            area_data = [["Area", "Count"]]
            area_data.extend([[item.get('area', ''), str(item.get('count', 0))] for item in area_distribution[:10]])
            area_table = Table(area_data)
            area_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(area_table)
            story.append(Spacer(1, 20))

        # Crime type distribution table
        crime_type_distribution = data.get('crime_type_distribution', [])
        if crime_type_distribution:
            story.append(Paragraph("Crimes by Type (Top 10):", styles['Heading2']))
            type_data = [["Crime Type", "Count"]]
            type_data.extend([[item.get('crime_type', ''), str(item.get('count', 0))] for item in crime_type_distribution[:10]])
            type_table = Table(type_data)
            type_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(type_table)

    try:
        doc.build(story)
    except Exception as e:
        logger.error(f"Error generating system health PDF: {e}")
    return filepath


def generate_crime_summary_excel(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate Excel report for crime summary."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crime_summary_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Crime Summary"

        # Title
        ws['A1'] = "Crime Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 4

        # Filters
        if filters:
            ws.cell(row=row, column=1, value="Filters Applied:").font = Font(bold=True)
            row += 1
            for key, value in filters.items():
                if value:
                    ws.cell(row=row, column=1, value=f"{key}: {value}")
                    row += 1
            row += 1

        # Summary statistics
        ws.cell(row=row, column=1, value="Summary Statistics:").font = Font(bold=True)
        row += 1

        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row += 1

        summary = data.get('summary', {})
        summary_data = [
            ("Total Crimes", summary.get('total_crimes', 0)),
            ("Unique Areas", summary.get('unique_areas', 0)),
            ("Unique Crime Types", summary.get('unique_crime_types', 0)),
            ("High Risk %", f"{summary.get('high_risk_percentage', 0.0):.1f}%"),
            ("Medium Risk %", f"{summary.get('medium_risk_percentage', 0.0):.1f}%"),
            ("Low Risk %", f"{summary.get('low_risk_percentage', 0.0):.1f}%"),
        ]

        for metric, value in summary_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Area distribution
        area_distribution = data.get('area_distribution', [])
        if area_distribution:
            ws.cell(row=row, column=1, value="Crimes by Area (Top 10):").font = Font(bold=True)
            row += 1

            headers = ["Area", "Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            row += 1

            for item in area_distribution[:10]:
                ws.cell(row=row, column=1, value=item.get('area', ''))
                ws.cell(row=row, column=2, value=item.get('count', 0))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column if column[0].column is not None else 1)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    return filepath


def generate_user_activity_pdf(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate PDF report for user activity."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"user_activity_{timestamp}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
    )
    story.append(Paragraph("User Activity Report", title_style))
    story.append(Spacer(1, 12))

    # Report info
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 12))

    # User statistics
    story.append(Paragraph("User Statistics:", styles['Heading2']))
    stats = data.get('user_statistics', {})
    stats_data = [
        ["Metric", "Count"],
        ["Total Users", str(stats.get('total_users', 0))],
        ["Regular Users", str(stats.get('regular_users', 0))],
        ["Admin Users", str(stats.get('admin_users', 0))],
        ["Super Admin Users", str(stats.get('superadmin_users', 0))],
        ["New Today", str(stats.get('new_today', 0))],
        ["New This Week", str(stats.get('new_this_week', 0))],
        ["New This Month", str(stats.get('new_this_month', 0))],
    ]

    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 20))

    # Activity summary
    activity_summary = data.get('activity_summary', [])
    if activity_summary:
        story.append(Paragraph("Activity Summary:", styles['Heading2']))
        activity_data = [["Activity Type", "Count"]]
        activity_data.extend([[item.get('activity_type', ''), str(item.get('count', 0))] for item in activity_summary])

        activity_table = Table(activity_data)
        activity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(activity_table)

    doc.build(story)
    return filepath


def generate_system_health_pdf(data: Dict[str, Any], report_style: str = "table") -> str:
    """Generate PDF report for system health."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"system_health_{timestamp}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
    )
    story.append(Paragraph("System Health Report", title_style))
    story.append(Spacer(1, 12))

    # Report info
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 12))

    # System metrics
    story.append(Paragraph("System Metrics:", styles['Heading2']))
    metrics = data.get('system_metrics', {})
    metrics_data = [
        ["Metric", "Value"],
        ["Active Users", str(metrics.get('active_users', 0))],
        ["Total Crimes", str(metrics.get('total_crimes', 0))],
        ["Recent Activities (24h)", str(metrics.get('recent_activities_24h', 0))],
        ["ML Model Status", metrics.get('ml_model_status', 'Unknown')],
    ]

    metrics_table = Table(metrics_data)
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))

    try:
        doc.build(story)
    except Exception as e:
        logger.error(f"Error generating system health PDF: {e}")
    return filepath




def get_reports_from_db(limit: int = 50) -> List[Dict[str, Any]]:
    """Get reports from database."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            """
            SELECT id, title, type, format, file_path, file_size, status, created_by, created_at, parameters
            FROM reports
            ORDER BY created_at DESC
            LIMIT %s
            """,
            (limit,)
        )

        reports = []
        rows = cursor.fetchall()
        for row in rows:
            # Cast row to Dict for type checking since cursor is set to dictionary=True
            row_data = cast(Dict[str, Any], row)
            created_at = row_data.get("created_at")
            reports.append({
                "id": row_data.get("id"),
                "title": row_data.get("title"),
                "type": row_data.get("type"),
                "created_at": created_at.isoformat() if created_at and hasattr(created_at, 'isoformat') else None,
                "status": row_data.get("status"),
                "format": row_data.get("format"),
                "file_size": row_data.get("file_size") or 0,
                "file_path": row_data.get("file_path")
            })

        return reports

    except Exception as e:
        logger.error(f"Error getting reports from DB: {e}")
        return []
    finally:
        cursor.close()
        conn.close()


def generate_user_activity_excel(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate Excel report for user activity."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"user_activity_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "User Activity"

        # Title
        ws['A1'] = "User Activity Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 4

        # Filters
        if filters:
            ws.cell(row=row, column=1, value="Filters Applied:").font = Font(bold=True)
            row += 1
            for key, value in filters.items():
                if value:
                    ws.cell(row=row, column=1, value=f"{key}: {value}")
                    row += 1
            row += 1

        # User statistics
        ws.cell(row=row, column=1, value="User Statistics:").font = Font(bold=True)
        row += 1

        headers = ["Metric", "Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row += 1

        stats = data.get('user_statistics', {})
        stats_data = [
            ("Total Users", stats.get('total_users', 0)),
            ("Regular Users", stats.get('regular_users', 0)),
            ("Admin Users", stats.get('admin_users', 0)),
            ("Super Admin Users", stats.get('superadmin_users', 0)),
            ("New Today", stats.get('new_today', 0)),
            ("New This Week", stats.get('new_this_week', 0)),
            ("New This Month", stats.get('new_this_month', 0)),
        ]

        for metric, value in stats_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Activity summary
        activity_summary = data.get('activity_summary', [])
        if activity_summary:
            ws.cell(row=row, column=1, value="Activity Summary:").font = Font(bold=True)
            row += 1

            headers = ["Activity Type", "Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            row += 1

            for item in activity_summary:
                ws.cell(row=row, column=1, value=item.get('activity_type', ''))
                ws.cell(row=row, column=2, value=item.get('count', 0))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column if column[0].column is not None else 1)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    return filepath


def generate_system_health_excel(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate Excel report for system health."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"system_health_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "System Health"

        # Title
        ws['A1'] = "System Health Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        row = 4

        # System metrics
        ws.cell(row=row, column=1, value="System Metrics:").font = Font(bold=True)
        row += 1

        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row += 1

        metrics = data.get('system_metrics', {})
        metrics_data = [
            ("Active Users", metrics.get('active_users', 0)),
            ("Total Crimes", metrics.get('total_crimes', 0)),
            ("Recent Activities (24h)", metrics.get('recent_activities_24h', 0)),
            ("ML Model Status", metrics.get('ml_model_status', 'Unknown')),
        ]

        for metric, value in metrics_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Database health
        ws.cell(row=row, column=1, value="Database Health:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=data.get('database_health', 'Unknown'))
        row += 2

        # Table statistics
        table_stats = data.get('table_statistics', {})
        if table_stats:
            ws.cell(row=row, column=1, value="Table Statistics:").font = Font(bold=True)
            row += 1

            headers = ["Table", "Record Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            row += 1

            for table, count in table_stats.items():
                ws.cell(row=row, column=1, value=table)
                ws.cell(row=row, column=2, value=str(count))
                row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column if column[0].column is not None else 1)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    return filepath


def generate_crime_summary_csv(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate CSV report for crime summary."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crime_summary_{timestamp}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        csvfile.write("Crime Summary Report\n")
        csvfile.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # Filters applied
        if filters:
            csvfile.write("Filters Applied:\n")
            for key, value in filters.items():
                if value:
                    csvfile.write(f"{key}: {value}\n")
            csvfile.write("\n")

        # Summary statistics
        csvfile.write("Summary Statistics:\n")
        csvfile.write("Metric,Value\n")
        summary = data.get('summary', {})
        csvfile.write(f"Total Crimes,{summary.get('total_crimes', 0)}\n")
        csvfile.write(f"Unique Areas,{summary.get('unique_areas', 0)}\n")
        csvfile.write(f"Unique Crime Types,{summary.get('unique_crime_types', 0)}\n")
        csvfile.write(f"High Risk %,{summary.get('high_risk_percentage', 0.0):.1f}%\n")
        csvfile.write(f"Medium Risk %,{summary.get('medium_risk_percentage', 0.0):.1f}%\n")
        csvfile.write(f"Low Risk %,{summary.get('low_risk_percentage', 0.0):.1f}%\n\n")

        # Area distribution
        area_distribution = data.get('area_distribution', [])
        if area_distribution:
            csvfile.write("Crimes by Area (Top 10):\n")
            csvfile.write("Area,Count\n")
            for item in area_distribution[:10]:
                csvfile.write(f"{item.get('area', '')},{item.get('count', 0)}\n")
            csvfile.write("\n")

        # Crime type distribution
        crime_type_distribution = data.get('crime_type_distribution', [])
        if crime_type_distribution:
            csvfile.write("Crimes by Type (Top 10):\n")
            csvfile.write("Crime Type,Count\n")
            for item in crime_type_distribution[:10]:
                csvfile.write(f"{item.get('crime_type', '')},{item.get('count', 0)}\n")

    return filepath


def generate_user_activity_csv(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate CSV report for user activity."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"user_activity_{timestamp}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        csvfile.write("User Activity Report\n")
        csvfile.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # User statistics
        csvfile.write("User Statistics:\n")
        csvfile.write("Metric,Count\n")
        stats = data.get('user_statistics', {})
        csvfile.write(f"Total Users,{stats.get('total_users', 0)}\n")
        csvfile.write(f"Regular Users,{stats.get('regular_users', 0)}\n")
        csvfile.write(f"Admin Users,{stats.get('admin_users', 0)}\n")
        csvfile.write(f"Super Admin Users,{stats.get('superadmin_users', 0)}\n")
        csvfile.write(f"New Today,{stats.get('new_today', 0)}\n")
        csvfile.write(f"New This Week,{stats.get('new_this_week', 0)}\n")
        csvfile.write(f"New This Month,{stats.get('new_this_month', 0)}\n\n")

        # Activity summary
        activity_summary = data.get('activity_summary', [])
        if activity_summary:
            csvfile.write("Activity Summary:\n")
            csvfile.write("Activity Type,Count\n")
            for item in activity_summary:
                csvfile.write(f"{item.get('activity_type', '')},{item.get('count', 0)}\n")

    return filepath


def generate_system_health_csv(data: Dict[str, Any], filters: Dict[str, Any], report_style: str = "table") -> str:
    """Generate CSV report for system health."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"system_health_{timestamp}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        csvfile.write("System Health Report\n")
        csvfile.write(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # System metrics
        csvfile.write("System Metrics:\n")
        csvfile.write("Metric,Value\n")
        metrics = data.get('system_metrics', {})
        csvfile.write(f"Active Users,{metrics.get('active_users', 0)}\n")
        csvfile.write(f"Total Crimes,{metrics.get('total_crimes', 0)}\n")
        csvfile.write(f"Recent Activities (24h),{metrics.get('recent_activities_24h', 0)}\n")
        csvfile.write(f"ML Model Status,{metrics.get('ml_model_status', 'Unknown')}\n\n")

        # Database health
        csvfile.write(f"Database Health:,{data.get('database_health', 'Unknown')}\n\n")

        # Table statistics
        table_stats = data.get('table_statistics', {})
        if table_stats:
            csvfile.write("Table Statistics:\n")
            csvfile.write("Table,Record Count\n")
            for table, count in table_stats.items():
                csvfile.write(f"{table},{count}\n")

    return filepath


def get_scheduled_reports_from_db(limit: int = 50) -> List[Dict[str, Any]]:
    """Get scheduled reports from database."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            """
            SELECT id, title, type, schedule, format, recipients, parameters, next_run, last_run, status, created_by, created_at
            FROM scheduled_reports
            WHERE status = 'active'
            ORDER BY next_run ASC
            LIMIT %s
            """,
            (limit,)
        )

        scheduled = []
        rows = cursor.fetchall()
        for row in rows:
            # Cast row to Dict for type checking since cursor is set to dictionary=True
            row_data = cast(Dict[str, Any], row)
            recipients = json.loads(row_data.get("recipients") or "[]")
            next_run = row_data.get("next_run")
            scheduled.append({
                "id": row_data.get("id"),
                "title": row_data.get("title"),
                "schedule": row_data.get("schedule"),
                "next_run": next_run.isoformat() if next_run and hasattr(next_run, 'isoformat') else None,
                "format": row_data.get("format"),
                "recipients": recipients,
                "status": row_data.get("status")
            })

        return scheduled

    except Exception as e:
        logger.error(f"Error getting scheduled reports from DB: {e}")
        return []
    finally:
        cursor.close()
        conn.close()


# Helper functions for report data generation (without FastAPI dependencies)
def get_crime_summary_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    area: Optional[str] = None,
    crime_type: Optional[str] = None,
    report_style: Optional[str] = None
) -> Dict[str, Any]:
    """Get crime summary report data without FastAPI dependencies."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Build query with filters
        query = """
            SELECT
                COUNT(*) as total_crimes,
                COUNT(DISTINCT area) as unique_areas,
                COUNT(DISTINCT crime_type) as unique_crime_types,
                AVG(CASE WHEN risk_level = 'High' THEN 1 ELSE 0 END) * 100 as high_risk_percentage,
                AVG(CASE WHEN risk_level = 'Medium' THEN 1 ELSE 0 END) * 100 as medium_risk_percentage,
                AVG(CASE WHEN risk_level = 'Low' THEN 1 ELSE 0 END) * 100 as low_risk_percentage
            FROM crimes
            WHERE 1=1
        """
        params = []

        if start_date:
            query += " AND crime_date >= %s"
            params.append(start_date)
        if end_date:
            query += " AND crime_date <= %s"
            params.append(end_date)
        if area:
            query += " AND area = %s"
            params.append(area)
        if crime_type:
            query += " AND crime_type = %s"
            params.append(crime_type)

        cursor.execute(query, params)
        summary = cursor.fetchone()

        # If no crimes match the filter, return empty summary
        if not summary or (isinstance(summary, dict) and summary.get("total_crimes", 0) == 0):
            return {
                "summary": {"total_crimes": 0},
                "area_distribution": [],
                "crime_type_distribution": [],
                "monthly_trend": [],
                "filters_applied": {
                    "start_date": start_date,
                    "end_date": end_date,
                    "area": area,
                    "crime_type": crime_type
                }
            }

        # Get crime distribution by area
        area_query = """
            SELECT area, COUNT(*) as crime_count
            FROM crimes
            WHERE 1=1
        """
        area_params = []
        if start_date:
            area_query += " AND crime_date >= %s"
            area_params.append(start_date)
        if end_date:
            area_query += " AND crime_date <= %s"
            area_params.append(end_date)
        if crime_type:
            area_query += " AND crime_type = %s"
            area_params.append(crime_type)

        area_query += " GROUP BY area ORDER BY crime_count DESC LIMIT 10"
        cursor.execute(area_query, area_params)
        area_distribution = cursor.fetchall()

        # Get crime distribution by type
        type_query = """
            SELECT crime_type, COUNT(*) as crime_count
            FROM crimes
            WHERE 1=1
        """
        type_params = []
        if start_date:
            type_query += " AND crime_date >= %s"
            type_params.append(start_date)
        if end_date:
            type_query += " AND crime_date <= %s"
            type_params.append(end_date)
        if area:
            type_query += " AND area = %s"
            type_params.append(area)

        type_query += " GROUP BY crime_type ORDER BY crime_count DESC LIMIT 10"
        cursor.execute(type_query, type_params)
        type_distribution = cursor.fetchall()

        # Get monthly trend (last 12 months, filtered by date range if provided)
        trend_query = """
            SELECT
                DATE_FORMAT(crime_date, '%Y-%m') as month,
                COUNT(*) as crime_count
            FROM crimes
            WHERE 1=1
        """
        trend_params = []
        if start_date:
            trend_query += " AND crime_date >= %s"
            trend_params.append(start_date)
        if end_date:
            trend_query += " AND crime_date <= %s"
            trend_params.append(end_date)
        if area:
            trend_query += " AND area = %s"
            trend_params.append(area)
        if crime_type:
            trend_query += " AND crime_type = %s"
            trend_params.append(crime_type)

        trend_query += " GROUP BY DATE_FORMAT(crime_date, '%Y-%m') ORDER BY month"
        cursor.execute(trend_query, trend_params)
        monthly_trend = cursor.fetchall()

        return {
            "summary": {
                "total_crimes": cast(Dict[str, Any], summary)["total_crimes"],
                "unique_areas": cast(Dict[str, Any], summary)["unique_areas"],
                "unique_crime_types": cast(Dict[str, Any], summary)["unique_crime_types"],
                "high_risk_percentage": round(cast(Dict[str, Any], summary)["high_risk_percentage"] or 0, 2),
                "medium_risk_percentage": round(cast(Dict[str, Any], summary)["medium_risk_percentage"] or 0, 2),
                "low_risk_percentage": round(cast(Dict[str, Any], summary)["low_risk_percentage"] or 0, 2)
            },
            "area_distribution": [
                {"area": cast(Dict[str, Any], row)["area"], "count": cast(Dict[str, Any], row)["crime_count"]}
                for row in area_distribution
            ],
            "crime_type_distribution": [
                {"crime_type": cast(Dict[str, Any], row)["crime_type"], "count": cast(Dict[str, Any], row)["crime_count"]}
                for row in type_distribution
            ],
            "monthly_trend": [
                {"month": cast(Dict[str, Any], row)["month"], "count": cast(Dict[str, Any], row)["crime_count"]}
                for row in monthly_trend
            ],
            "filters_applied": {
                "start_date": start_date,
                "end_date": end_date,
                "area": area,
                "crime_type": crime_type
            }
        }

    except Exception as e:
        logger.error(f"Database error generating crime summary report: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def get_user_activity_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_role: Optional[str] = None,
    report_style: Optional[str] = None
) -> Dict[str, Any]:
    """Get user activity report data without FastAPI dependencies."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # User registration statistics
        reg_query = """
            SELECT
                COUNT(*) as total_users,
                COUNT(CASE WHEN role = 'user' THEN 1 END) as regular_users,
                COUNT(CASE WHEN role = 'admin' THEN 1 END) as admin_users,
                COUNT(CASE WHEN role = 'superadmin' THEN 1 END) as superadmin_users,
                COUNT(CASE WHEN DATE(created_at) = CURDATE() THEN 1 END) as new_today,
                COUNT(CASE WHEN created_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY) THEN 1 END) as new_this_week,
                COUNT(CASE WHEN created_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN 1 END) as new_this_month
            FROM users_info
            WHERE 1=1
        """
        reg_params = []

        if user_role:
            reg_query += " AND role = %s"
            reg_params.append(user_role)

        cursor.execute(reg_query, reg_params)
        user_stats = cursor.fetchone()

        # User activity logs summary
        activity_query = """
            SELECT
                activity_type,
                COUNT(*) as count
            FROM user_activity_logs
            WHERE 1=1
        """
        activity_params = []

        if start_date:
            activity_query += " AND DATE(created_at) >= %s"
            activity_params.append(start_date)
        if end_date:
            activity_query += " AND DATE(created_at) <= %s"
            activity_params.append(end_date)

        activity_query += " GROUP BY activity_type ORDER BY count DESC"
        cursor.execute(activity_query, activity_params)
        activity_summary = cursor.fetchall()

        # Recent user registrations
        recent_query = """
            SELECT username, first_name, last_name, email, role, created_at
            FROM users_info
            WHERE 1=1
        """
        recent_params = []

        if user_role:
            recent_query += " AND role = %s"
            recent_params.append(user_role)

        recent_query += " ORDER BY created_at DESC LIMIT 20"
        cursor.execute(recent_query, recent_params)
        recent_users = cursor.fetchall()

        return {
            "user_statistics": {
                "total_users": cast(Dict[str, Any], user_stats)["total_users"],
                "regular_users": cast(Dict[str, Any], user_stats)["regular_users"],
                "admin_users": cast(Dict[str, Any], user_stats)["admin_users"],
                "superadmin_users": cast(Dict[str, Any], user_stats)["superadmin_users"],
                "new_today": cast(Dict[str, Any], user_stats)["new_today"],
                "new_this_week": cast(Dict[str, Any], user_stats)["new_this_week"],
                "new_this_month": cast(Dict[str, Any], user_stats)["new_this_month"]
            },
            "activity_summary": [
                {"activity_type": cast(Dict[str, Any], row)["activity_type"], "count": cast(Dict[str, Any], row)["count"]}
                for row in activity_summary
            ],
            "recent_registrations": [
                {
                    "username": cast(Dict[str, Any], row)["username"],
                    "name": f"{cast(Dict[str, Any], row)['first_name']} {cast(Dict[str, Any], row)['last_name']}",
                    "email": cast(Dict[str, Any], row)["email"],
                    "role": cast(Dict[str, Any], row)["role"],
                    "created_at": cast(Dict[str, Any], row)["created_at"].isoformat() if cast(Dict[str, Any], row)["created_at"] else None
                }
                for row in recent_users
            ],
            "filters_applied": {
                "start_date": start_date,
                "end_date": end_date,
                "user_role": user_role
            }
        }

    except Exception as e:
        logger.error(f"Database error generating user activity report: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def get_system_health_data(report_style: Optional[str] = None) -> Dict[str, Any]:
    """Get system health report data without FastAPI dependencies."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Database connection health
        db_health = "healthy"
        try:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        except Exception as e:
            db_health = f"unhealthy: {str(e)}"

        # Table statistics
        tables_stats = {}
        tables = ["users_info", "crimes", "admins", "user_activity_logs", "audit_logs"]

        for table in tables:
            try:
                cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
                result = cursor.fetchone()
                tables_stats[table] = cast(Dict[str, Any], result)["count"]
            except Exception as e:
                tables_stats[table] = f"error: {str(e)}"

        # System metrics
        cursor.execute("SELECT COUNT(*) as active_users FROM users_info WHERE role = 'user'")
        active_users_result = cursor.fetchone()
        active_users = cast(Dict[str, Any], active_users_result)["active_users"]

        cursor.execute("SELECT COUNT(*) as total_crimes FROM crimes")
        total_crimes_result = cursor.fetchone()
        total_crimes = cast(Dict[str, Any], total_crimes_result)["total_crimes"]

        cursor.execute("SELECT COUNT(*) as recent_activities FROM user_activity_logs WHERE created_at >= DATE_SUB(NOW(), INTERVAL 24 HOUR)")
        recent_activities_result = cursor.fetchone()
        recent_activities = cast(Dict[str, Any], recent_activities_result)["recent_activities"]

        # ML model status - check if model files exist in the expected location
        try:
            from .core.config import MODEL_DIR
            import os
            import joblib

            model_path = os.path.join(MODEL_DIR, 'random_forest_model.joblib')
            le_area_path = os.path.join(MODEL_DIR, 'label_encoder_area.joblib')
            le_crime_path = os.path.join(MODEL_DIR, 'label_encoder_crime.joblib')
            le_risk_path = os.path.join(MODEL_DIR, 'label_encoder_risk.joblib')

            ml_status = "loaded" if all(os.path.exists(p) for p in [model_path, le_area_path, le_crime_path, le_risk_path]) else "not_loaded"
        except Exception as e:
            logger.warning(f"Could not check ML model status: {e}")
            ml_status = "unknown"

        return {
            "database_health": db_health,
            "table_statistics": tables_stats,
            "system_metrics": {
                "active_users": active_users,
                "total_crimes": total_crimes,
                "recent_activities_24h": recent_activities,
                "ml_model_status": ml_status
            },
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        logger.error(f"Database error generating system health report: {e}")
        raise
    finally:
        cursor.close()
        conn.close()


def download_report(report_id: int) -> str:
    """Download a report by report_id and return the file path.

    Args:
        report_id: The ID of the report to download

    Returns:
        The file path of the report

    Raises:
        ValueError: If the report is not found or file doesn't exist
    """
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Get report from database
        cursor.execute(
            "SELECT id, file_path FROM reports WHERE id = %s",
            (report_id,)
        )
        report = cursor.fetchone()

        if not report:
            raise ValueError(f"Report with id {report_id} not found")

        file_path = cast(Dict[str, Any], report).get("file_path")
        if not file_path:
            raise ValueError(f"Report {report_id} has no file path")

        # Normalize path for Windows
        normalized_path = os.path.normpath(file_path)

        if not os.path.exists(normalized_path):
            raise ValueError(f"Report file not found at {normalized_path}")

        logger.info(f"Report {report_id} found at {normalized_path}")
        return normalized_path

    except Exception as e:
        logger.error(f"Error downloading report {report_id}: {e}")
        raise
    finally:
        cursor.close()
        conn.close()
