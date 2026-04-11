# app/utils/report_generator.py
"""
Professional Report Generation Module
Supports PDF, Excel, CSV, and JSON formats with charts and visualizations
"""

import os
import json
import csv
from datetime import datetime
from typing import Dict, List, Any, Optional
from io import BytesIO
import base64

# PDF Generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF generation will be limited.")

# Excel Generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel generation will be limited.")

# Chart Generation
try:
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Warning: matplotlib not installed. Chart generation will be limited.")


class ReportGenerator:
    """Professional report generator with multiple format support"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
    def generate_report(
        self,
        report_type: str,
        data: Dict[str, Any],
        format: str,
        filters: Dict[str, Any],
        filename: str
    ) -> str:
        """
        Generate a report in the specified format
        
        Args:
            report_type: Type of report (crime_summary, user_activity, system_health)
            data: Report data
            format: Output format (pdf, excel, csv, json)
            filters: Report filters and options
            filename: Output filename
            
        Returns:
            Path to generated file
        """
        filepath = os.path.join(self.output_dir, filename)
        
        if format == 'pdf':
            return self._generate_pdf(report_type, data, filters, filepath)
        elif format == 'excel':
            return self._generate_excel(report_type, data, filters, filepath)
        elif format == 'csv':
            return self._generate_csv(report_type, data, filters, filepath)
        elif format == 'json':
            return self._generate_json(report_type, data, filters, filepath)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_pdf(self, report_type: str, data: Dict, filters: Dict, filepath: str) -> str:
        """Generate professional PDF report with charts"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation")
        
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a3a5f'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#00a6a6'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        title = self._get_report_title(report_type)
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Report metadata
        metadata = [
            ['Report Type:', report_type.replace('_', ' ').title()],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Period:', f"{filters.get('start_date', 'N/A')} to {filters.get('end_date', 'N/A')}"],
            ['Format:', 'PDF Document']
        ]
        
        metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
        metadata_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        story.append(metadata_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Summary Section
        if 'summary' in data and filters.get('includeStatistics', True):
            story.append(Paragraph("Executive Summary", heading_style))
            summary_data = self._format_summary_for_pdf(data['summary'])
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(self._get_table_style())
            story.append(summary_table)
            story.append(Spacer(1, 0.2*inch))
        
        # Statistical Analysis Section
        if filters.get('includeStatistics', True):
            story.append(PageBreak())
            story.append(Paragraph("Statistical Analysis", heading_style))
            
            # Crime Types Distribution
            if 'crime_types' in data and data['crime_types']:
                story.append(Paragraph("Crime Type Distribution", styles['Heading3']))
                crime_type_data = [['Crime Type', 'Count', 'Percentage']]
                total = sum(item['count'] for item in data['crime_types'])
                for item in data['crime_types']:
                    percentage = (item['count'] / total * 100) if total > 0 else 0
                    crime_type_data.append([
                        item['crime_type'],
                        str(item['count']),
                        f"{percentage:.1f}%"
                    ])
                
                crime_table = Table(crime_type_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
                crime_table.setStyle(self._get_table_style())
                story.append(crime_table)
                story.append(Spacer(1, 0.2*inch))
            
            # Trend Analysis
            if 'trends' in data and data['trends']:
                story.append(Paragraph("Trend Analysis", styles['Heading3']))
                trend_summary = f"Data collected over {len(data['trends'])} days in the selected period."
                story.append(Paragraph(trend_summary, styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
        
        # Charts (if requested)
        if filters.get('includeCharts', True) and MATPLOTLIB_AVAILABLE:
            story.append(PageBreak())
            story.append(Paragraph("Data Visualizations", heading_style))
            
            # Generate charts based on report type
            if report_type == 'crime_summary' and 'top_areas' in data and data['top_areas']:
                chart_path = self._create_bar_chart(data['top_areas'], 'Crime Distribution by Area')
                if chart_path:
                    img = Image(chart_path, width=5*inch, height=3*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
            
            # Crime types pie chart
            if report_type == 'crime_summary' and 'crime_types' in data and data['crime_types']:
                pie_chart_path = self._create_pie_chart(data['crime_types'], 'Crime Types Distribution')
                if pie_chart_path:
                    img = Image(pie_chart_path, width=5*inch, height=3*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2*inch))
        
        # Detailed Data Tables
        if filters.get('includeRawData', False):
            story.append(PageBreak())
            story.append(Paragraph("Detailed Data", heading_style))
            
            if 'top_areas' in data and data['top_areas']:
                area_data = [['Area', 'Incident Count']]
                for area in data['top_areas']:
                    area_data.append([area['area'], str(area['count'])])
                
                area_table = Table(area_data, colWidths=[4*inch, 2*inch])
                area_table.setStyle(self._get_table_style())
                story.append(area_table)
                story.append(Spacer(1, 0.2*inch))
            
            if 'crime_types' in data and data['crime_types']:
                crime_data = [['Crime Type', 'Count']]
                for crime in data['crime_types']:
                    crime_data.append([crime['crime_type'], str(crime['count'])])
                
                crime_table = Table(crime_data, colWidths=[4*inch, 2*inch])
                crime_table.setStyle(self._get_table_style())
                story.append(crime_table)
        
        # Build PDF
        doc.build(story)
        return filepath
    
    def _generate_excel(self, report_type: str, data: Dict, filters: Dict, filepath: str) -> str:
        """Generate professional Excel report with charts"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Summary"
        
        # Styling
        header_fill = PatternFill(start_color="1a3a5f", end_color="1a3a5f", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(size=16, bold=True, color="1a3a5f")
        
        # Title
        ws['A1'] = self._get_report_title(report_type)
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = 'Report Type:'
        ws[f'B{row}'] = report_type.replace('_', ' ').title()
        row += 1
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        ws[f'A{row}'] = 'Period:'
        ws[f'B{row}'] = f"{filters.get('start_date', 'N/A')} to {filters.get('end_date', 'N/A')}"
        row += 2
        
        # Summary Section
        if 'summary' in data:
            ws[f'A{row}'] = 'Executive Summary'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="00a6a6")
            row += 1
            
            for key, value in data['summary'].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            row += 1
        
        # Top Areas Data
        if 'top_areas' in data and data['top_areas']:
            ws[f'A{row}'] = 'Top Areas by Incident Count'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="00a6a6")
            row += 1
            
            # Headers
            ws[f'A{row}'] = 'Area'
            ws[f'B{row}'] = 'Incident Count'
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            row += 1
            
            # Data
            chart_start_row = row
            for area in data['top_areas']:
                ws[f'A{row}'] = area['area']
                ws[f'B{row}'] = area['count']
                row += 1
            
            # Add Chart if requested
            if filters.get('includeCharts', True):
                chart = BarChart()
                chart.title = "Incident Distribution by Area"
                chart.style = 10
                chart.y_axis.title = 'Incident Count'
                chart.x_axis.title = 'Area'
                
                data_ref = Reference(ws, min_col=2, min_row=chart_start_row-1, max_row=row-1)
                cats = Reference(ws, min_col=1, min_row=chart_start_row, max_row=row-1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f'D{chart_start_row}')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    if column_letter is None:
                        column_letter = cell.column_letter
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    def _generate_csv(self, report_type: str, data: Dict, filters: Dict, filepath: str) -> str:
        """Generate CSV report"""
        # utf-8-sig keeps CSV readable in Excel while preserving Unicode characters.
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header
            writer.writerow([self._get_report_title(report_type)])
            writer.writerow([])
            writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow(['Period:', f"{filters.get('start_date', 'N/A')} to {filters.get('end_date', 'N/A')}"])
            writer.writerow([])
            
            # Summary
            if 'summary' in data:
                writer.writerow(['Executive Summary'])
                for key, value in data['summary'].items():
                    writer.writerow([key.replace('_', ' ').title(), value])
                writer.writerow([])
            
            # Top Areas
            if 'top_areas' in data and data['top_areas']:
                writer.writerow(['Top Areas by Incident Count'])
                writer.writerow(['Area', 'Incident Count'])
                for area in data['top_areas']:
                    writer.writerow([area['area'], area['count']])
                writer.writerow([])

            # Crime types
            if 'crime_types' in data and data['crime_types']:
                writer.writerow(['Crime Types'])
                writer.writerow(['Crime Type', 'Count'])
                for crime in data['crime_types']:
                    writer.writerow([crime.get('crime_type'), crime.get('count')])
                writer.writerow([])

            # Activity types (user activity reports)
            if 'activities' in data and data['activities']:
                writer.writerow(['Activity Types'])
                writer.writerow(['Activity Type', 'Count'])
                for activity in data['activities']:
                    writer.writerow([activity.get('activity_type'), activity.get('count')])
                writer.writerow([])

            # Trends
            if 'trends' in data and data['trends']:
                writer.writerow(['Trend Data'])
                writer.writerow(['Date', 'Count', 'High Risk'])
                for trend in data['trends']:
                    writer.writerow([
                        trend.get('date'),
                        trend.get('count'),
                        trend.get('high_risk', ''),
                    ])
        
        return filepath
    
    def _generate_json(self, report_type: str, data: Dict, filters: Dict, filepath: str) -> str:
        """Generate JSON report"""
        from decimal import Decimal
        from datetime import date, datetime as dt
        
        def decimal_default(obj):
            """JSON serializer for objects not serializable by default json code"""
            if isinstance(obj, Decimal):
                return float(obj)
            if isinstance(obj, (date, dt)):
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
        
        report = {
            'metadata': {
                'report_type': report_type,
                'title': self._get_report_title(report_type),
                'generated_at': datetime.now().isoformat(),
                'period': {
                    'start_date': filters.get('start_date'),
                    'end_date': filters.get('end_date')
                },
                'filters': filters
            },
            'data': data
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(report, jsonfile, indent=2, ensure_ascii=False, default=decimal_default)
        
        return filepath
    
    def _create_bar_chart(self, data: List[Dict], title: str) -> Optional[str]:
        """Create a bar chart using matplotlib"""
        if not MATPLOTLIB_AVAILABLE or not data:
            return None
        
        try:
            plt.figure(figsize=(10, 6))
            areas = [item['area'] for item in data[:10]]
            counts = [item['count'] for item in data[:10]]
            
            plt.barh(areas, counts, color='#00a6a6')
            plt.xlabel('Incident Count')
            plt.title(title)
            plt.tight_layout()
            
            chart_path = os.path.join(self.output_dir, f'chart_{datetime.now().timestamp()}.png')
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
        except Exception as e:
            print(f"Error creating chart: {e}")
            return None
    
    def _create_pie_chart(self, data: List[Dict], title: str) -> Optional[str]:
        """Create a pie chart using matplotlib"""
        if not MATPLOTLIB_AVAILABLE or not data:
            return None
        
        try:
            plt.figure(figsize=(10, 6))
            labels = [item['crime_type'] for item in data[:8]]  # Top 8
            sizes = [item['count'] for item in data[:8]]
            
            colors = ['#00a6a6', '#1a3a5f', '#f59e0b', '#dc2626', '#10b981', '#8b5cf6', '#ec4899', '#f97316']
            
            plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors, startangle=90)
            plt.title(title)
            plt.axis('equal')
            
            chart_path = os.path.join(self.output_dir, f'pie_chart_{datetime.now().timestamp()}.png')
            plt.savefig(chart_path, dpi=150, bbox_inches='tight')
            plt.close()
            
            return chart_path
        except Exception as e:
            print(f"Error creating pie chart: {e}")
            return None
    
    def _get_report_title(self, report_type: str) -> str:
        """Get formatted report title"""
        titles = {
            'crime_summary': 'Crime Incident Summary Report',
            'user_activity': 'User Activity Analysis Report',
            'system_health': 'System Health & Performance Report'
        }
        return titles.get(report_type, 'System Report')
    
    def _format_summary_for_pdf(self, summary: Dict) -> List[List[str]]:
        """Format summary data for PDF table"""
        data = [['Metric', 'Value']]
        for key, value in summary.items():
            formatted_key = key.replace('_', ' ').title()
            data.append([formatted_key, str(value)])
        return data
    
    def _get_table_style(self) -> TableStyle:
        """Get standard table style for PDF"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ])
